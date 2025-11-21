"""
Row Loss Diagnostic Tool
Investigates why rows are being lost during Excel processing.

Usage: python diagnose_row_loss.py <path_to_excel_file>
"""

import sys
import pandas as pd
import numpy as np


def diagnose_file(file_path):
    """
    Comprehensive diagnosis of Excel file reading issues.
    """
    print("=" * 70)
    print("ROW LOSS DIAGNOSTIC TOOL")
    print("=" * 70)
    print(f"\nFile: {file_path}\n")

    # ========== Test 1: Default pandas read ==========
    print("TEST 1: Default pandas.read_excel()")
    print("-" * 70)
    try:
        df1 = pd.read_excel(file_path, engine='openpyxl')
        print(f"✓ Rows read: {df1.shape[0]}")
        print(f"  Columns: {df1.shape[1]}")
        na_rows = df1.isna().all(axis=1).sum()
        print(f"  Completely NA rows: {na_rows}")
    except Exception as e:
        print(f"✗ Error: {e}")

    # ========== Test 2: With keep_default_na=False ==========
    print("\nTEST 2: With keep_default_na=False")
    print("-" * 70)
    try:
        df2 = pd.read_excel(file_path, engine='openpyxl', keep_default_na=False)
        print(f"✓ Rows read: {df2.shape[0]}")
        print(f"  Columns: {df2.shape[1]}")
        empty_rows = (df2 == '').all(axis=1).sum()
        print(f"  Empty string rows: {empty_rows}")
    except Exception as e:
        print(f"✗ Error: {e}")

    # ========== Test 3: With dtype=str ==========
    print("\nTEST 3: With dtype=str")
    print("-" * 70)
    try:
        df3 = pd.read_excel(file_path, engine='openpyxl', dtype=str)
        print(f"✓ Rows read: {df3.shape[0]}")
        print(f"  Columns: {df3.shape[1]}")
    except Exception as e:
        print(f"✗ Error: {e}")

    # ========== Test 4: Current production settings ==========
    print("\nTEST 4: Production settings (keep_default_na=False, dtype=str)")
    print("-" * 70)
    try:
        df4 = pd.read_excel(
            file_path,
            engine='openpyxl',
            keep_default_na=False,
            dtype=str
        )
        print(f"✓ Rows read: {df4.shape[0]}")
        print(f"  Columns: {df4.shape[1]}")
    except Exception as e:
        print(f"✗ Error: {e}")

    # ========== Test 5: NEW - With na_filter=False ==========
    print("\nTEST 5: UPDATED settings (keep_default_na=False, dtype=str, na_filter=False)")
    print("-" * 70)
    try:
        df5 = pd.read_excel(
            file_path,
            engine='openpyxl',
            keep_default_na=False,
            dtype=str,
            na_filter=False
        )
        print(f"✓ Rows read: {df5.shape[0]}")
        print(f"  Columns: {df5.shape[1]}")
    except Exception as e:
        print(f"✗ Error: {e}")

    # ========== Test 6: Direct openpyxl read ==========
    print("\nTEST 6: Direct openpyxl.load_workbook()")
    print("-" * 70)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Count non-empty rows (excluding header)
        non_empty = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                non_empty += 1

        print(f"✓ Non-empty rows (excluding header): {non_empty}")
        print(f"  Total rows in sheet: {ws.max_row - 1}")  # -1 for header
        print(f"  Sheet name: {ws.title}")
    except Exception as e:
        print(f"✗ Error: {e}")

    # ========== Comparison Analysis ==========
    print("\n" + "=" * 70)
    print("ANALYSIS")
    print("=" * 70)

    try:
        counts = []
        labels = []

        if 'df1' in locals():
            counts.append(df1.shape[0])
            labels.append("Default read")

        if 'df2' in locals():
            counts.append(df2.shape[0])
            labels.append("keep_default_na=False")

        if 'df3' in locals():
            counts.append(df3.shape[0])
            labels.append("dtype=str")

        if 'df4' in locals():
            counts.append(df4.shape[0])
            labels.append("Production settings")

        if 'df5' in locals():
            counts.append(df5.shape[0])
            labels.append("Updated settings")

        if 'non_empty' in locals():
            counts.append(non_empty)
            labels.append("Direct openpyxl")

        max_count = max(counts)
        max_label = labels[counts.index(max_count)]

        print(f"\n📊 Row Count Summary:")
        for label, count in zip(labels, counts):
            diff = max_count - count
            status = "✓" if count == max_count else f"✗ Missing {diff} rows"
            print(f"  {label:30s}: {count:4d} rows  {status}")

        print(f"\n💡 Recommendation:")
        print(f"  Use: {max_label}")
        print(f"  This captures the most rows ({max_count})")

        if max_count != df4.shape[0]:
            print(f"\n⚠️  WARNING: Production settings losing {max_count - df4.shape[0]} rows!")
            print(f"  Update excel_utils.py to use the recommended settings.")

    except Exception as e:
        print(f"Error during analysis: {e}")

    # ========== Column Analysis ==========
    print("\n" + "=" * 70)
    print("COLUMN ANALYSIS")
    print("=" * 70)

    try:
        if 'df4' in locals():
            print(f"\nColumns in file:")
            for i, col in enumerate(df4.columns, 1):
                print(f"  {i:2d}. {col}")

            if 'text' in df4.columns:
                print(f"\n'text' column analysis:")
                text_col = df4['text'].fillna('').astype(str)
                print(f"  Non-empty: {(text_col.str.strip() != '').sum()}")
                print(f"  Empty/NA: {(text_col.str.strip() == '').sum()}")
            else:
                print(f"\n⚠️  WARNING: No 'text' column found!")
                print(f"  Available columns: {', '.join(df4.columns)}")

    except Exception as e:
        print(f"Error during column analysis: {e}")

    print("\n" + "=" * 70)
    print("DIAGNOSIS COMPLETE")
    print("=" * 70)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python diagnose_row_loss.py <path_to_excel_file>")
        sys.exit(1)

    file_path = sys.argv[1]
    diagnose_file(file_path)